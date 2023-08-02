import openpyxl
import re
import requests

# Excel Function definitions

def SUM(*args):
    if not all(isinstance(arg, int) for arg in args):
        return ("Arguments must be integer values")
    return sum(args)


def MULTIPLY(*args):
    result = 1
    for  arg in args:
        result *= arg
    return result

def DIVIDE(Dividend, Divisor):
    result = Dividend/Divisor
    return result

def GT(first, second):
    return(first> second)

def EQ(first,second):
    return first == second

def NOT(a):
    return not a

def AND(*args):
    for arg in args:
        if not isinstance(arg, bool):
            return "#ERROR: Incompatible types"
    return all(args)

def OR(*args):
    for arg in args:
        if not isinstance(arg, bool):
            return "#ERROR: Incompatible types"
    return any(args)

def CONCAT(*args):
    result = ""
    for a in args:
         result+= a
    return a

# def if_error(form, stringone, worksheet ):
#     try: 
#         formula_to_value(form,worksheet)
#     except:
#         return(strig)

def  extract_formula(worksheet,cell):
    formula = str(cell.value)
    return formula

def is_it_if_boi(formula):
    return "IF(" in formula



def split_formula(formula): #function specially made to split the function if it is made with the IF formula
        if "IFERROR" in formula:
            operands = formula[9:len(formula)-1]
        else:
            operands = formula[4:len(formula)-1]
        
        parts = []
        level = 0
        start = 0
        for i in range(len(operands)):
            if operands[i] == '(' :
                level += 1
            elif operands[i] == ')':
                level -= 1
            elif operands[i] == ',' and level == 0:
                parts.append(operands[start:i])
                start = i + 1
        parts.append(operands[start:])
        return parts



def split_formula_for_concat(formula): #function specially made to split the function if it is made with the concat formula
    # remove the function name and opening parenthesis
    a = formula.index('(')
    args_str = formula[a+1:-1]

    # split the arguments using commas
    args_list = []
    quote_count = 0
    start = 0
    for i in range(len(args_str)):
        if args_str[i] == ',' and quote_count % 2 == 0:
            args_list.append(args_str[start:i])
            start = i + 1
        elif args_str[i] == '"':
            quote_count += 1

    # append the final argument to the list
    args_list.append(args_str[start:])

    # remove any extra whitespace and quotes from the arguments
    args_list = [arg.strip().strip('"') for arg in args_list]

    return args_list

# here Operator is the "if" part in =if("GT(A1,B1)","Yes","No")

# need a function that handles 

def formula_to_parts(formula):

    
    
    indexOpening = formula.index('(')
    indexClosing = formula.index(')')

    if formula.startswith('=') :
        operator = formula[1:indexOpening]
    else:    
        operator = formula[:indexOpening]

    raw = formula[indexOpening+1: indexClosing]

    operands = [uno.strip() for uno in raw.split(', ')]

    return operator,operands

def is_cell_reference(s):
    pattern = r"^[A-Z]+\d+$"
    return bool(re.match(pattern, s))

def value_them_refs(operands,worksheet):
    for key, ref in enumerate(operands):
        if is_cell_reference(ref): # to check if value is cell reference
                cell_ref = ref
                cell_value = worksheet[cell_ref].value
                operands[key] = cell_value
        elif "true" in operands[key]:
                operands[key] = True
        elif "false" in operands[key]:
                operands[key] = False
        elif ref.isdigit():
             operands[key] =int(ref)
             
        
    return operands

def calculate_equation(operator,operands):
     
    if operator == "SUM":
        result = SUM(*operands)
        return result
    elif operator == "MULTIPLY":
            result = MULTIPLY(*operands)
            return result
    elif operator == "DIVIDE" :
        result = DIVIDE(operands[0],operands[1])
        return result
    elif operator == "GT" :
        result = GT(operands[0],operands[1])
        return result
    elif operator == "EQ":
        result = EQ(operands[0],operands[1])
        return result
    elif operator == "NOT":
        result = NOT(*operands)
        return result

    elif operator == "AND":
            result = AND(*operands)
            return result

    elif operator == "OR":
        result = OR(*operands)
        return result
    
    elif operator == "CONCAT":
         result = CONCAT(*operands)
         return result
    
def formula_to_value(formula,worksheet):
    if "(" in formula:  
        operator,operands = formula_to_parts(formula)
        
        operands = value_them_refs(operands,worksheet)
        result = calculate_equation(operator,operands)
        return result
        
    else:
        operands = []
        operands.append(formula[1:])
        a = value_them_refs(operands,worksheet)
        newstr = a[0]
        if isinstance(newstr,str) and newstr.startswith('='):
            return formula_to_value(newstr,worksheet)
        else:
            return newstr
    

        


def eval_sheet(data):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in data:
        worksheet.append(row)

    workbook.save('activesheet.xlsx')

    workbook = openpyxl.load_workbook('activesheet.xlsx')
    worksheet = workbook.active

    listtoreturn = []

    for row in worksheet.iter_rows():
        
        listinsidelist = []
        for cell in row:
            
            if cell.data_type =='f' : # to check if it is a string or not
                
                    
                formula = extract_formula(worksheet,cell)
                
                if is_it_if_boi(formula):
                    operands = split_formula(formula)
                    newformula = operands[0]
                    if formula_to_value(newformula,worksheet):
                        values = []
                        values.append(operands[1].strip())
                        a = value_them_refs(values,worksheet)
                        result = a[0]
                    else:
                        values = []
                        values.append(operands[2].strip())
                        a = value_them_refs(values,worksheet)
                        result = a[0]

                elif "CONCAT" in formula:
                    operands = split_formula_for_concat(formula)
                    operands = value_them_refs(operands,worksheet)
                    result = ""
                    for a in operands:
                        result+=a

                elif "IFERROR" in formula:
                    operands = split_formula(formula)
                    newform = operands[0]
                    #print(formula, " Hellloooooooooooooooooooooo", newform, operands)      ************************
                    try: 
                        
                        a = formula_to_value(newform,worksheet)
                        #print(newform, a)                                                     *********************
                        result = a
                    except:
                        #print(newform, operands[1])                                           **********************
                        result = operands[1]


                
                else:
                    result = formula_to_value(formula,worksheet)


            else:
                formula = extract_formula(worksheet,cell)                
                if formula.startswith('='):
                    result = formula_to_value(formula,worksheet)
                else:
                    result = cell.value
            
            listinsidelist.append(result)
        listtoreturn.append(listinsidelist)
    return listtoreturn

# -----------------------------------------------------------Code for Validation 

response = requests.get('https://www.wix.com/_serverless/hiring-task-spreadsheet-evaluator/sheets')

# Parse response JSON
response_data = response.json()

# Get submission URL and sheets data
submission_url = response_data['submissionUrl']
sheets_data = response_data['sheets']

results = []
for sheet in sheets_data:
    
# Process sheets data and generate results
        sheet_id = sheet['id']
        sheet_data = sheet['data']
        
        
        # Evaluating sheet data here and generate result
        
        result = {'id': sheet_id, 'data': eval_sheet(sheet_data)}
        
        # Appending result to list of results
        results.append(result)
    
        #print(sheet_id, sheet_data)                            *********************

# # Make POST request to submission URL with results data
payload = {'email': 'haridethekd@gmail.com', 'results': results}
#print(results, "resultsssss")                                  
response = requests.post(submission_url, json=payload)

# Parse response JSON
response_data = response.json()



# Get passcode from response
print(response_data)



#passcode 1dff47768745cd6f1a341529dc70a12d